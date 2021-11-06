import re
import csv
from openpyxl import Workbook

def read_phonebook(file_name):
    with open(file_name, encoding="utf-8") as file:
        items = csv.reader(file, delimiter=",")
        contacts_list = list(items)
    return contacts_list

def changed_phone_number_to_new(contacts_list):
    numbers_origin = r'(\+7|8)(\s*)(\(*)(\d{3})(\)*)(\s*)' \
                            r'(\-*)(\d{3})(\s*)(\-*)(\d{2})(\s*)(\-*)' \
                            r'(\d{2})(\s*)(\(*)(доб)*(\.*)(\s*)(\d+)*(\)*)'
    new_number = r'+7(\4)\8-\11-\14\15\17\18\19\20'
    file_rows_number_updated = []
    for row in contacts_list:
        rows_strings = ','.join(row)
        final_rows_here = re.sub(numbers_origin, new_number, rows_strings)
        rows_now_list = final_rows_here.split(',')
        file_rows_number_updated.append(rows_now_list)
    return file_rows_number_updated

def changed_full_name_to_change_to_new(contacts_list):
    names_original = r'^([А-ЯЁа-яё]+)(\s*)(\,?)([А-ЯЁа-яё]+)' \
                       r'(\s*)(\,?)([А-ЯЁа-яё]*)(\,?)(\,?)(\,?)'
    names_now_edited = r'\1\3\10\4\6\9\7\8'
    file_rows_names_updated = []
    for row in contacts_list:
        rows_strings = ','.join(row)
        final_rows_here = re.sub(names_original, names_now_edited, rows_strings)
        rows_now_list = final_rows_here.split(',')
        file_rows_names_updated.append(rows_now_list)
    return file_rows_names_updated

def Join_it_all_together(contacts_list):
    for rows_verion_1 in contacts_list:
        for rows_verion_2 in contacts_list:
            if rows_verion_1[0] == rows_verion_2[0] and rows_verion_1[1] ==\
                    rows_verion_2[1] and rows_verion_1 is not rows_verion_2:
                if rows_verion_1[2] == ' ':
                    rows_verion_1[2] = rows_verion_2[2]
                if rows_verion_1[3] == ' ':
                    rows_verion_1[3] = rows_verion_2[3]
                if rows_verion_1[4] == ' ':
                    rows_verion_1[4] = rows_verion_2[4]
                if rows_verion_1[5] == ' ':
                    rows_verion_1[5] = rows_verion_2[5]
                if rows_verion_1[6] == ' ':
                    rows_verion_1[6] = rows_verion_2[6]
    rows_updated_joined_together = []
    for row in contacts_list:
        if row not in rows_updated_joined_together:
            rows_updated_joined_together.append(row)
    return rows_updated_joined_together

def write_file_of_excel(contacts_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Phone Book"
    for item in range(8):
        ws.append(contacts_list[item])
    ws['G8'].value = contacts_list[8][6]
    ws['E3'].value = ws['E5'].value
    ws.delete_rows(5)
    wb.save('phonebook_updated.xlsx')


if __name__ == '__main__':
    contacts = read_phonebook('phonebook_raw.csv')
    contacts = changed_phone_number_to_new(contacts)
    contacts = changed_full_name_to_change_to_new(contacts)
    contacts = Join_it_all_together(contacts)
    write_file_of_excel(contacts)

